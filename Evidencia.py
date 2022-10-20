# Estructura de datos y su procesamiento

# Evidencia

# Renta de espacios de coworking

# Imports

import csv
import datetime
import openpyxl

# Diccionarios

clientes_registrados = {}
nombres_salas = {}
eventos = {}

# Listas

fecha_turno = []

# Conjuntos

conjunto_universal = set()
conjunto_ocupados = set()

# Variables

dia = 1
dia_dos = 2
fecha_hoy = datetime.date.today()

# Menu

while True :

    print ("\n****************")
    print ("   BIENVENIDO   ")
    print ("****************")
    print ("\n MENU PRINCIPAL")
    print ("\n[1] Editar el nombre del evento de una reservación ya hecha")
    print ("\n[2] Consultar las reservaciones existentes")
    print ("\n[3] Registrase como nuevo cliente")
    print ("\n[4] Rentar una sala")
    print ("\n[5] Salir")

    menu_opcion = int (input ("\nPor favor escribe el numero del menu al que desea acceder : "))

    # Menu opcion 1

    if menu_opcion == 1 :
        print ("\n-- EDITAR EL NOMBRE DEL EVENTO DE UNA RESERVACION YA HECHA --")

        acceso_dos = int (input ("\nPor favor ingresa la clave de tu evento : "))

        # Acceso

        if acceso_dos in eventos :
            print ("\n-----------------")
            print ("Evento encontrado")
            print ("-----------------")

            # Editar evento

            print ("\nEl nombre de tu evento es : ", nombres_salas[acceso_dos])
            nuevo_nombre = input ("\nEscribe el nuevo nombre para tu evento : ")
            nombres_salas [acceso_dos] = nuevo_nombre

            sala = nuevo_nombre
            nombre_para_registro = eventos [acceso_dos] [4]
            turno = eventos [acceso_dos] [3]
            fecha = eventos [acceso_dos] [2]
            asistentes = eventos [acceso_dos] [1]

            eventos [acceso_dos] = sala, asistentes, fecha, turno, nombre_para_registro

            print ("\n-------------------------------------------")
            print ("El nombre del evento se edito correctamente")
            print ("\nEl nuevo nombre de tu evento es : ", nuevo_nombre)
            print ("-------------------------------------------")

        # Acceso

        else :
            print ("\n--------------------")
            print ("Evento no encontrado")
            print ("--------------------")

        regreso = int (input ('\nEscribe "0" para regresar al menu principal : '))

    # Menu opcion 2

    elif menu_opcion == 2 :
        print ("\n-- CONSULTAR LAS RESERVACIONES EXISTENTES --")

        print ("\n[1] Consultar las reservaciones de una fecha en especifico")
        print ("\n[2] Consultar las reservaciones disponibles de una fecha en especifico")

        opcion = int (input("\nPor favor escribe el numero del menu al que deseas acceder : "))

        # Reservaciones

        if opcion == 1 :
            print ('\n[1] Consultar las reservaciones en "pantalla"')
            print ('\n[2] Consultar las reservaciones en "excel"')

            opcion_dos = int (input("\nEscribe el numero del menu al que deseas acceder : "))

            # Reservaciones en pantalla

            if opcion_dos == 1 :
                print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
                consulta = input("\nEscribe la fecha en la que quiere consultar las reservaciones : ")
                consulta = datetime.datetime.strptime (consulta,'%d/%m/%Y').date()

                lista_reservas = list(eventos.items())

                print ("\n----------------------------------------------------------------")
                print (f"                   Reporte del dia {consulta}")
                print ("----------------------------------------------------------------")
                print ("Numero de sala       Cliente       Nombre del evento       Turno")
                print ("----------------------------------------------------------------")

                for clave,valor in lista_reservas :
                    if valor [2] == consulta :
                        print ("     ",clave,"             ",valor[4],"           ",valor[0],"           ",valor[3])
                        print ("----------------------------------------------------------------")

            # Reservaciones en excel

            elif opcion_dos == 2 :
                print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
                consulta_tres = input("\nEscribe la fecha en la que quiere consultar las reservaciones : ")
                consulta_tres = datetime.datetime.strptime (consulta_tres,'%d/%m/%Y').date()

                lista_reservas = list(eventos.items())

                libro = openpyxl.Workbook()
                hoja = libro["Sheet"]
                hoja.title = "Primera"

                hoja["B1"].value = "Reporte para el dia : "
                hoja["C1"].value = consulta_tres
                hoja["A2"].value = "Numero de sala"
                hoja["B2"].value = "Cliente"
                hoja["C2"].value = "Nombre del evento"
                hoja["D2"].value = "Turno"

                for clave,valor in lista_reservas :
                    if valor [2] == consulta_tres :
                        hoja.cell(row=clave,column=1).value = clave
                        hoja.cell(row=clave,column=2).value = valor[4]
                        hoja.cell(row=clave,column=3).value = valor[0]
                        hoja.cell(row=clave,column=4).value = valor[3]

                libro.save("MiExcelDesdePython.xlsx")

                print ("\n-----------------------------------------------------------")
                print ("Reporte creado exitosamente / Por favor revisa el documento")
                print ("-----------------------------------------------------------")

        # Reservaciones disponibles

        elif opcion == 2 :
            print ('\nEscribe la fecha con el siguiente formato "09/01/2004"')
            consulta_dos = input("\nEscribe la fecha en la que quieres consultar las reservaciones disponibles : ")
            consulta_dos = datetime.datetime.strptime (consulta_dos,'%d/%m/%Y').date()

            turnos_disponibles = conjunto_universal - conjunto_ocupados
            turnos_disponibles = list(turnos_disponibles)
            turnos_disponibles.sort()

            print ('\nLos turnos disponibles para el dia "', consulta_dos,'" son los siguientes\n')

            for clave_dos,valor_dos in turnos_disponibles :
                if clave_dos == consulta_dos :
                    print ("--------------------")
                    print ("Turno disponible :",valor_dos)
                    print ("--------------------")

        else :
            print ("\n---------------------------------")
            print ("El valor que ingreso no es valido")
            print ("---------------------------------")

        regreso = int (input ('\nEscribe "0" para regresar al menu principal : '))

    # Menu opcion 3

    elif menu_opcion == 3 :
        print ("\n-- REGISTRARSE COMO NUEVO CLIENTE --")

        while True :
            cliente_nuevo = input ("\nPor favor escribe tu nombre completo : ")

            if cliente_nuevo == "" :
                print ("\n------------------------------")
                print ("El nombre no puede ser omitido")
                print ("------------------------------")

            # Nuevo cliente

            else:
                llave = max (list (clientes_registrados.keys()), default=0) + 1
                clientes_registrados [llave] = cliente_nuevo
                print ("\n--------------------------------------------------------")
                print (f"Felicidades {cliente_nuevo} tu registro concluyo exitosamente")
                print (f"\nTu clave es : {llave}")
                print ("--------------------------------------------------------")
                break

        regreso = int (input ('\nEscribe "0" para regresar al menu principal : '))

    # Menu opcion 4

    elif menu_opcion == 4 :
        print ("\n-- RENTAR UNA SALA --")

        print ("\nPara rentar una sala es necesario ser cliente registrado")
        acceso = int (input("\nPor favor ingresa tu clave como cliente : "))

        # Validacion cliente

        while True :
            if acceso in clientes_registrados :
                print ("\n----------------------------------------------")
                print ("Cliente encontrado / Bienvenid@", clientes_registrados[acceso])
                print ("----------------------------------------------")

                nombre_para_registro = clientes_registrados[acceso]

                # Nombre evento

                while True :
                    sala = input ("\nEscribe el nombre del evento para registrar tu sala : ")

                    if sala == "" :
                        print ("\n------------------------------")
                        print ("El nombre no puede ser omitido")
                        print ("------------------------------")

                    else :
                        llave_dos = max (list(nombres_salas.keys()), default=3) + 1
                        nombres_salas [llave_dos] = sala
                        print ("\n--------------------------")
                        print(f"La clave de tu sala es : {llave_dos}")
                        print ("--------------------------")
                        break

                # Asistentes

                while True :
                    asistentes = int (input("\nEscribe la cantidad de asistentes para tu evento : "))

                    if asistentes == 0 :
                        print ("\n-------------------------------------------------")
                        print ('La cantidad de asistentes debe de ser mayor a "0"')
                        print ("-------------------------------------------------")

                    elif asistentes > 0 :
                        print ("\n----------------------")
                        print ("Asistentes registrados")
                        print ("----------------------")
                        break

                # Fecha

                while True :
                    print ("\nEscribre la fecha para tu evento con el siguiente formato (Dia/Mes/Año)")
                    print ('\nEjemplo : "09/01/2004"')
                    print ("\nLa fecha debe de ser con dos dias de anticipacion")

                    fecha = input ("\nEscribe la fecha para tu evento : ")
                    fecha = datetime.datetime.strptime (fecha,'%d/%m/%Y').date()

                    if fecha <= fecha_hoy :
                        print ("\n-------------------------------------------------------------------")
                        print ("Fecha no valida / La fecha debe de ser con dos dias de anticipacion")
                        print ("-------------------------------------------------------------------")


                    elif fecha == fecha_hoy + datetime.timedelta(days=+dia) :
                        print ("\n-------------------------------------------------------------------")
                        print ("Fecha no valida / La fecha debe de ser con dos dias de anticipacion")
                        print ("-------------------------------------------------------------------")


                    elif fecha >= fecha_hoy + datetime.timedelta(days=+dia_dos) :
                        print ("\n----------------")
                        print ("Fecha disponible")
                        print ("----------------")

                        # Horario

                        while True :

                            print ("\nHorarios disponibles")
                            print ("\n[1] Mañana")
                            print ("\n[2] Tarde")
                            print ("\n[3] Noche")

                            turno = int (input ("\nEscribe el turno en el que deseas registrar tu sala : "))
                            datos = [(fecha,turno)]

                            if datos[0] in fecha_turno :
                                print ("\n------------------------------------------------------------------")
                                print ("Turno no disponible / Ya existe un evento registrado en este turno")
                                print ("------------------------------------------------------------------")

                            elif turno >= 4 :
                                print ("\n------------------------------")
                                print ("El turno que ingreso no existe")
                                print ("------------------------------")

                            else :
                                print ("\n----------------")
                                print ("Turno disponible")
                                print ("----------------")
                                datos.remove((fecha,turno))
                                break
                        break


                # Registro finalizado

                print ("\nTu sala se ha registrado exitosamente")
                print ("\n-------------------------------------------------------------------")
                print("Nombre   \tClave \t    Cantidad \t    Fecha   \t      Turno")
                print (f"{sala}  \t  {llave_dos} \t      {asistentes}        {fecha}    \t{turno} ")
                print ("-------------------------------------------------------------------")

                fecha_turno.append((fecha,turno))

                eventos [llave_dos] = sala, asistentes, fecha, turno, nombre_para_registro

                conjunto_ocupados.add((fecha,turno))

                conjunto_universal.add((fecha,1))
                conjunto_universal.add((fecha,2))
                conjunto_universal.add((fecha,3))

                break

            # Validacion cliente

            else :
                print ("\n---------------------")
                print ("Cliente no encontrado")
                print ("---------------------")
                break

        regreso = int (input ('\nEscribe "0" para regresar al menu principal : '))

    # Menu opcion 5

    elif menu_opcion == 5 :
        print ("\n-- SALIDA --")

        print ("\n-------------------------")
        print ("El programa ha finalizado")
        print ("-------------------------")

        # CSV

        # clientes_registrados

        with open("clientes.csv","w", newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(("Clave", "nombre"))

            for clave,valor in clientes_registrados.items():
                registro = clave,valor
                grabador.writerow(registro)

        # nombres_salas

        with open("salas.csv","w", newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(("Clave", "nombre"))

            for clave,valor in nombres_salas.items():
                registro = clave,valor
                grabador.writerow(registro)

        # eventos

        with open("eventos.csv","w", newline="") as archivo:
            grabador = csv.writer(archivo)
            grabador.writerow(("Clave", "nombre"))

            for clave,valor in eventos.items():
                registro = clave,valor[0],valor[1],valor[2],valor[3],valor[4]
                grabador.writerow(registro)

        # fecha_turno

        with open("fechaturno.csv", "w", newline="") as archivo:
            grabador = csv.writer(archivo)

            for elemento in fecha_turno:
                grabador.writerow(elemento)

        # conjunto_ocupados

        with open("conjuntosocupados.cvs", "w", newline="") as archivo:
            grabador = csv.writer(archivo)

            for elemento in conjunto_ocupados:
                grabador.writerow(elemento)

        break

    # Menu dato no valido

    else :
        print ("\n------------------------------------------------------------------")
        print ("La opcion que escribio no es valida / Por favor intentalo de nuevo")
        print ("------------------------------------------------------------------")

        regreso = int (input ('\nEscribe "0" para regresar al menu principal : '))